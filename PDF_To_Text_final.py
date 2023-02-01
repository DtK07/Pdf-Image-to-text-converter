from PyPDF2 import PdfFileReader
from pathlib import Path
import openpyxl
from openpyxl import workbook, load_workbook
import fitz
import pytesseract
import PIL.Image

wb = openpyxl.load_workbook("Anna_Univ_Colleges.xlsx")
sh = wb['Colleges']
sh['A1'].value = "File_Name"
sh['B1'].value = "College_Name"
sh['C1'].value = "City"
sh['D1'].value = "Email_ID"
sh['E1'].value = "Phone"
sh['F1'].value = "Website"
sh['G1'].value = "Type"

input_dir = Path.cwd()/"Full list"
files  = list(input_dir.glob("*.pdf*"))
file_names =[]

for item in files:
    name = item.name
    file_names.append(name)
print(file_names)
print(len(file_names))

i=2
for name in file_names:
    j=1
    sh.cell(i, j).value = name
    j+=1
    Pdf = PdfFileReader(f"C:\\Users\\DINESH\\PycharmProjects\\PDF To Text\\Full list\\{name}")
    page_1_object = Pdf.getPage(0)
    page_1_text = page_1_object.extractText().strip()
    if page_1_text == "":
        pdf = fitz.open(f"C:\\Users\\DINESH\\PycharmProjects\\PDF To Text\\Full list\\{name}")
        img_name = name.replace(".pdf","")
        # print(pdf)
        image_list = pdf.get_page_images(0)
        # print(image_list)
        for image in image_list:
            xref = image[0]
            # print(xref)
            pix = fitz.Pixmap(pdf, xref)
            if pix.n < 5:
                pix.save(f"{img_name}.png")
            else:
                pix1 = fitz.open(fitz.csRGB, pix)
                pix1.save(f"{img_name}.png")
                pix1 = None
            pix = None
        # print(len(image_list))
        pytesseract.pytesseract.tesseract_cmd = r"C:\Users\DINESH\AppData\Local\Programs\Tesseract-OCR/tesseract.exe"
        myconfig = r"--psm 6 --oem 3"
        data = pytesseract.image_to_string(PIL.Image.open(f"{img_name}.png"), config=myconfig).strip()
        # print(data)
        C_Name = data[data.find(':') + 1: data.find('2.')].strip().split(",")
        name = C_Name[0]
        name1 = name.split(",")
        College_Name = name1[0]
        sh.cell(i, j).value = College_Name
        j += 1
        Address = data[data.find('Address') + len('Address') + 1: data.find('Name of the Principal') - 3]
        City = Address.split(",")[-1].replace(".", "").strip().strip(":").strip(
            ">").strip(":").strip("*").strip("_").strip()
        sh.cell(i, j).value = City
        j += 1
        Email = data[data.find('Email ID') + len('Email ID') + 1: data.find('Year of Establishment') - 3].strip().strip(":").strip(
            ">").strip(":").strip("*").strip("_").strip()
        sh.cell(i, j).value = Email
        j += 1
        Phone = data[data.find('Phone Number') + len('Phone Number') + 1: data.find('Fax Number') - 3].strip().strip(":").strip(
            ">").strip(":").strip("*").strip("_").strip()
        sh.cell(i, j).value = Phone
        j += 1
        Website = data[data.find('Web Site') + len('Web Site') + 1: data.find('Email ID') - 3].strip().strip(":").strip(
            ">").strip(":").strip("*").strip("_").strip()
        sh.cell(i, j).value = Website
        j+=1
        Type = data[data.find('Type') + len('Type') + 1: data.find('Minority Status') - 4].strip().strip(":").strip(
            ">").strip(":").strip("*").strip("_").strip()
        sh.cell(i, j).value = Type
        j += 1
        i += 1

    else:
        College_Name = page_1_text[page_1_text.find('Name of the College')+len('Name of the College')+1: page_1_text.find('Address')-3]
        sh.cell(i, j).value = College_Name
        j+=1
        Address = page_1_text[page_1_text.find('Address')+len('Address')+1: page_1_text.find('Name of the Principal')]
        City = Address.split(",")[-1].replace(".","")
        sh.cell(i, j).value = City
        j+=1
        Email = page_1_text[page_1_text.find('Email ID')+len('Email ID')+1: page_1_text.find('Year of Establishment')-3]
        sh.cell(i, j).value = Email
        j+=1
        Phone = page_1_text[page_1_text.find('Phone Number')+len('Phone Number')+1: page_1_text.find('Fax Number')-3]
        sh.cell(i, j).value = Phone
        j+=1
        Website = page_1_text[page_1_text.find('Web Site')+len('Web Site')+1: page_1_text.find('Email ID')-3]
        sh.cell(i, j).value = Website
        j+=1
        Type = page_1_text[page_1_text.find('Type')+len('Type')+1: page_1_text.find('Minority Status')-4]
        sh.cell(i, j).value = Type
        j+=1
        i+=1
wb.save("Anna_Univ_Colleges.xlsx")




